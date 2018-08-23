import openpyxl
import os
from rhymeCoordExtractor import rhymeTable, analyzeSentence


def initEmptyDB(filename):
    if not os.path.exists(filename):
        empdb = openpyxl.Workbook(filename)
    else:
        empdb = openpyxl.load_workbook(filename)
    for i in range(len(rhymeTable)):
        for j in range(len(rhymeTable)):
            if '('+str(i)+', '+str(j)+')' in empdb.sheetnames:
                continue
            else :
                empdb.create_sheet('('+str(i)+', '+str(j)+')')
    empdb.save(filename)
    empdb.close()


def analyzeAFile(filename,database):
    fin = open(filename,"r",encoding='utf-8')
    db = openpyxl.load_workbook(database)
    curS = fin.readline()
    print("current sentence = ", curS)
    while curS:
        curScoordRes, curSoriRes = analyzeSentence(curS)
        print("current result = ",curScoordRes,curSoriRes)
        for i in range(len(curScoordRes)):
            print(str(curScoordRes[i][-1]))
            cursheet = db[str(curScoordRes[i][-1])]
            row = [curSoriRes[i]]
            for m in range(len(curScoordRes[i])-1):
                row.append(str(curScoordRes[i][-2-m]))
            cursheet.append(row)
        curS = fin.readline()
        db.save(database)
    db.close()
    fin.close()


